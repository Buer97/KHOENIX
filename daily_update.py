# -*- coding: utf-8 -*-
"""
每日A股情绪数据自动更新脚本（云部署版）
- GitHub Actions 定时运行：每周一至五 15:35（北京时间）
- 从 akshare 抓取涨停/炸板/跌停/成交额数据
- 更新 Excel 并提交回仓库
"""
import sys
import os
from datetime import datetime, timezone, timedelta

# 云端路径（相对路径，Excel 存在仓库根目录）
EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "情绪周期.xlsx")

# 北京时间
BJT = timezone(timedelta(hours=8))
now_bjt = datetime.now(BJT)
date_str = now_bjt.strftime("%Y%m%d")

def log(msg):
    ts = now_bjt.strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")

# 确保依赖
for pkg in ["akshare", "openpyxl"]:
    try:
        __import__(pkg)
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

import akshare as ak
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)

def fetch_today_data(date_str):
    """抓取今日涨停/炸板/跌停/成交额数据"""
    result = {
        "zt_total": 0, "zt_count": 0, "zbgc_count": 0,
        "dt_total": 0, "dt_count": 0,
        "lianban": 0, "shouban": 0, "erban": 0,
        "sanban": "", "siban": "", "wuban": "",
        "liuban": "", "qiban": "", "baban": "",
        "jiuban": "", "shiban": "", "shiyiban": "",
        "volume_wan_yi": 0,
    }

    # 涨停池
    try:
        df_zt = ak.stock_zt_pool_em(date=date_str)
        if df_zt is not None and not df_zt.empty:
            result["zt_total"] = len(df_zt)
            result["zt_count"] = len(df_zt)
            lb_col = "连板数" if "连板数" in df_zt.columns else "连板天数"
            if lb_col in df_zt.columns:
                for n in range(1, 12):
                    stocks = df_zt[df_zt[lb_col] == n]["名称"].tolist()
                    if n == 1:
                        result["shouban"] = len(stocks)
                    elif n == 2:
                        result["erban"] = len(stocks)
                    else:
                        keys = ["sanban","siban","wuban","liuban","qiban","baban","jiuban","shiban","shiyiban"]
                        result[keys[n-3]] = "、".join(stocks[:3]) if stocks else ""
            lianban = df_zt[df_zt[lb_col] >= 2] if lb_col in df_zt.columns else []
            result["lianban"] = len(lianban)
        log(f"涨停池: {result['zt_total']}家")
    except Exception as e:
        log(f"涨停池失败: {e}")

    # 炸板池
    try:
        df_zbgc = ak.stock_zt_pool_zbgc_em(date=date_str)
        if df_zbgc is not None and not df_zbgc.empty:
            result["zbgc_count"] = len(df_zbgc)
        log(f"炸板池: {result['zbgc_count']}家")
    except Exception as e:
        log(f"炸板池失败: {e}")

    # 跌停池
    try:
        df_dt = ak.stock_zt_pool_dtgc_em(date=date_str)
        if df_dt is not None and not df_dt.empty:
            result["dt_total"] = len(df_dt)
            result["dt_count"] = len(df_dt)
        log(f"跌停池: {result['dt_total']}家")
    except Exception as e:
        log(f"跌停池失败: {e}")

    # 两市成交额（万亿）
    try:
        df_sh = ak.stock_zh_index_daily_em(symbol="sh000001")
        df_sz = ak.stock_zh_index_daily_em(symbol="sz399001")
        dt_target = pd.Timestamp(datetime.strptime(date_str, "%Y%m%d"))
        sh_row = df_sh[pd.to_datetime(df_sh['date']) == dt_target]
        sz_row = df_sz[pd.to_datetime(df_sz['date']) == dt_target]
        if not sh_row.empty and not sz_row.empty:
            total = sh_row['amount'].values[0] + sz_row['amount'].values[0]
            result["volume_wan_yi"] = round(total / 1e12, 2)
        log(f"成交额: {result['volume_wan_yi']}万亿")
    except Exception as e:
        log(f"成交额失败: {e}")

    denom = result["zt_count"] + result["zbgc_count"]
    result["zbgc_rate"] = round(result["zbgc_count"] / denom * 100, 1) if denom > 0 else 0.0
    return result

def append_to_excel(date_str, data):
    """将今日数据插入到表头下方第4行（最新的在最上面）"""
    wb = load_workbook(EXCEL_PATH)
    ws = wb["情绪周期"]

    dt = datetime.strptime(date_str, "%Y%m%d")
    weekday = ["一","二","三","四","五","六","日"][dt.weekday()]
    date_display = f"{dt.strftime('%Y/%m/%d')}({weekday})"

    # 幂等检查
    for row in ws.iter_rows(min_row=4, max_col=1, values_only=True):
        if row[0] and dt.strftime("%Y/%m/%d") in str(row[0]):
            log(f"{date_display} 数据已存在，跳过写入")
            wb.close()
            return False

    # 在第4行插入新行
    ws.insert_rows(4)
    insert_row = 4

    zbgc_rate = data["zbgc_rate"]
    row_vals = [
        date_display,
        "",                    # 盘口（手动）
        "",                    # 市场描述（手动）
        "",                    # 接力模式（手动）
        data["zt_total"],
        data["zt_count"],
        zbgc_rate,
        data["dt_total"],
        data["dt_count"],
        data["lianban"],
        data["shouban"],
        data["erban"],
        data["sanban"],
        data["siban"],
        data["wuban"],
        data["liuban"],
        data["qiban"],
        data["baban"],
        data["jiuban"],
        data["shiban"],
        data["shiyiban"],
        data["volume_wan_yi"],  # Col 22: 成交量(万亿)
        "",
    ]

    ws.row_dimensions[insert_row].height = 32
    for col_idx, val in enumerate(row_vals, 1):
        c = ws.cell(row=insert_row, column=col_idx, value=val)
        c.font = Font(size=14, name="微软雅黑")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()

        if col_idx == 1:
            c.fill = hex_fill("EAF0FF")
            c.font = Font(bold=True, color="2B4590", size=14, name="微软雅黑")
        elif col_idx == 2:
            c.fill = hex_fill("FFFDE7")
            c.font = Font(italic=True, color="999999", size=12, name="微软雅黑")
        elif col_idx == 7:
            if isinstance(val, (int, float)):
                if val >= 40:
                    c.fill = hex_fill("FFCDD2")
                    c.font = Font(bold=True, color="B71C1C", size=14, name="微软雅黑")
                elif val >= 25:
                    c.fill = hex_fill("FFF9C4")
                    c.font = Font(bold=True, color="F57F17", size=14, name="微软雅黑")
                else:
                    c.fill = hex_fill("E8F5E9")
                    c.font = Font(bold=True, color="1B5E20", size=14, name="微软雅黑")
        elif col_idx in (5, 6):
            c.fill = hex_fill("FFEBEE")
            c.font = Font(bold=True, color="CC0000", size=14, name="微软雅黑")
        elif col_idx in (8, 9):
            c.fill = hex_fill("E8F5E9")
            c.font = Font(bold=True, color="0A6D2E", size=14, name="微软雅黑")
        elif col_idx in (10, 11, 12):
            c.fill = hex_fill("F3E5F5")
            c.font = Font(bold=True, color="6A1B9A", size=14, name="微软雅黑")
        elif col_idx >= 13:
            c.fill = hex_fill("FFF8E1") if insert_row % 2 == 0 else hex_fill("FFFDF0")
            c.font = Font(color="5D4037", size=11, name="微软雅黑")
        else:
            c.fill = hex_fill("F8F9FD") if insert_row % 2 == 0 else hex_fill("FFFFFF")

    wb.save(EXCEL_PATH)
    log(f"✅ 写入 {date_display}: 涨停{data['zt_total']} 炸板率{zbgc_rate}% 跌停{data['dt_total']} 成交额{data['volume_wan_yi']}万亿")
    return True

# ─── 主逻辑 ──────────────────────────────────────────────────
if __name__ == "__main__":
    # 检查是否工作日
    if now_bjt.weekday() >= 5:
        log(f"今天是周末，非交易日，跳过")
        sys.exit(0)

    log(f"开始更新 {date_str} 的情绪数据...")

    if not os.path.exists(EXCEL_PATH):
        log(f"文件不存在: {EXCEL_PATH}")
        sys.exit(1)

    data = fetch_today_data(date_str)
    result = append_to_excel(date_str, data)

    if result:
        log("更新完成！")
    else:
        log("无需更新（数据已存在或无新数据）")
    log("脚本执行结束")
